#!/usr/bin/env python3
"""
海外空调客诉周报生成器
根据客诉台账数据，按周生成三个模块：本周新增 / 正常跟进 / 本周结案
"""
import os, sys
from datetime import datetime, timezone, timedelta, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

BJT = timezone(timedelta(hours=8))

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "2026年海外客户投诉台账.xlsx")

# 列映射 (台账列 → 周报列)
REPORT_COLS_ORDER = [
    "编号_int", "分公司", "国家或地区", "问题描述",
    "投诉日期", "应结案日期", "结案预警", "原因分析", "长期整改措施", "备注",
]
REPORT_COL_WIDTHS = [6, 6, 8, 50, 12, 12, 12, 30, 30, 20]
REPORT_HEADERS = [
    "编号", "分公司", "国家或地区", "详  细  描  述", "投诉日期",
    "应结案日期", "结案预警\n（最后三工作日）", "原因分析", "整改措施", "备注",
]


def load_data() -> pd.DataFrame:
    df = pd.read_excel(EXCEL_PATH, sheet_name="所有客诉", header=1)
    df["_num"] = pd.to_numeric(df["编号"], errors="coerce")
    df = df[df["_num"].notna() & df["分公司"].notna() & (df["分公司"].astype(str).str.strip() != "")].copy()
    df["编号_int"] = df["_num"].astype(int)
    for col in ["投诉日期", "应结案日期", "实际完成日期", "结案预警"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def format_date(d):
    if pd.isna(d) or d is None:
        return ""
    return d.strftime("%Y-%m-%d") if hasattr(d, 'strftime') else str(d)[:10]


def build_records(df: pd.DataFrame) -> list[dict]:
    """将 DataFrame 转为周报行列表"""
    rows = []
    for _, r in df.iterrows():
        row = {}
        for col_key in REPORT_COLS_ORDER:
            val = r.get(col_key, "")
            if col_key in ("编号_int",):
                try:
                    val = int(val) if pd.notna(val) and val != "" else ""
                except (ValueError, TypeError):
                    val = ""
            elif col_key in ("投诉日期", "应结案日期", "结案预警"):
                val = format_date(val) if pd.notna(val) else ""
            elif pd.isna(val) or val == "":
                val = ""
            else:
                val = str(val)
            row[col_key] = val
        rows.append(row)
    return rows


def get_week_dates(year: int, week: int) -> tuple:
    """返回 ISO 周的第一天和最后一天"""
    jan4 = date(year, 1, 4)
    start = jan4 - timedelta(days=jan4.isoweekday() - 1) + timedelta(weeks=week - 1)
    end = start + timedelta(days=6)
    return start, end


def generate_weekly_report(df: pd.DataFrame, year: int, week: int) -> Workbook:
    """生成周报 Excel"""
    start, end = get_week_dates(year, week)
    end_dt = pd.Timestamp(end)
    start_dt = pd.Timestamp(start)

    # ---- 本周新增 ----
    new_mask = (df["投诉日期"] >= start_dt) & (df["投诉日期"] <= end_dt)
    new_df = df[new_mask].sort_values("投诉日期")

    # ---- 正常跟进 ----
    follow_mask = (
        (df["投诉日期"] < start_dt) &
        df["应结案日期"].notna() &
        (df["应结案日期"] >= start_dt)
    )
    follow_df = df[follow_mask].sort_values("应结案日期")

    # ---- 本周结案 ----
    closed_mask = (
        df["实际完成日期"].notna() &
        (df["实际完成日期"] >= start_dt) &
        (df["实际完成日期"] <= end_dt)
    )
    closed_df = df[closed_mask].sort_values("实际完成日期")

    wb = Workbook()

    # ============================================
    # Sheet 1: 本周客诉
    # ============================================
    ws = wb.active
    ws.title = "本周客诉"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # 设置列宽
    for i, w in enumerate(REPORT_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num = 1

    # 标题行
    title = f"海 外 空 调 客 诉 第 {week} 周 周 报"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REPORT_HEADERS))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="微软雅黑", size=16, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    row_num = 2

    # 表头行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=1, value="项目\n状态").font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=2, column=1).alignment = center_align
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).border = thin_border

    for ci, h in enumerate(REPORT_HEADERS, 3):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
    ws.row_dimensions[2].height = 40
    row_num = 3

    def write_section(ws, start_row, section_name, records, col_count):
        """写入一个模块"""
        r = start_row
        # 合并第一列显示模块名
        if len(records) > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r + len(records) - 1, end_column=1)
        cell = ws.cell(row=r, column=1, value=section_name)
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = center_align
        cell.fill = section_fill

        for i, rec in enumerate(records):
            rr = r + i
            # 序号
            cell = ws.cell(row=rr, column=2, value=i + 1)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = center_align

            for ci, col_key in enumerate(REPORT_COLS_ORDER, 3):
                val = rec.get(col_key, "")
                cell = ws.cell(row=rr, column=ci, value=val)
                cell.font = Font(name="微软雅黑", size=9)
                cell.alignment = wrap_align

            # 边框
            for ci in range(1, col_count + 1):
                ws.cell(row=rr, column=ci).border = thin_border
            ws.row_dimensions[rr].height = max(60, 15 * len(str(rec.get("问题描述", ""))) // 60)

        # 第一列的合并单元格边框
        for i in range(len(records)):
            for ci in range(1, col_count + 1):
                ws.cell(row=r + i, column=ci).border = thin_border

        return r + max(len(records), 1)  # at least one row for empty section

    new_records = build_records(new_df)
    follow_records = build_records(follow_df)
    closed_records = build_records(closed_df)

    col_count = len(REPORT_HEADERS) + 1  # +1 for section column

    # 写入三个模块
    row_num = write_section(ws, row_num, "本周新增", new_records, col_count)
    row_num = write_section(ws, row_num, "正常跟进", follow_records, col_count)
    write_section(ws, row_num, "本周结案", closed_records, col_count)

    # ============================================
    # Sheet 2: 汇总
    # ============================================
    ws2 = wb.create_sheet("汇总")
    total_all = len(df)
    total_new = len(new_df)
    total_follow = len(follow_df)
    total_closed = len(closed_df)

    summary_data = [
        ["汇总表", "", "", "", "", ""],
        ["年份", "本周新增客诉（件）", "累计发生客诉（件）", "累计客服自行处理客诉（件）", "超期未结案客诉（件）", "同期累计已支付费用（元）"],
        [str(year), total_new, total_all, 0, 0, ""],
    ]

    # Row 1: title (write to cell before merge)
    ws2.cell(row=1, column=1, value="汇总表").font = Font(name="微软雅黑", size=14, bold=True)
    ws2.cell(row=1, column=1).alignment = center_align
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    for r, row_data in enumerate(summary_data[1:], 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 2:
                cell.font = Font(name="微软雅黑", size=10, bold=True)
                cell.fill = header_fill
            else:
                cell.font = Font(name="微软雅黑", size=10)
            cell.border = thin_border
            cell.alignment = center_align

    for i, w in enumerate([12, 20, 20, 25, 20, 25], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ============================================
    # Sheet 3: 情况通报
    # ============================================
    ws3 = wb.create_sheet("情况通报")
    ws3.column_dimensions['A'].width = 100
    ws3.sheet_properties.pageSetUpPr = None  # 允许自由文本布局

    bullet_font = Font(name="微软雅黑", size=11)
    title_font = Font(name="微软雅黑", size=14, bold=True)
    section_font = Font(name="微软雅黑", size=12, bold=True)
    body_font = Font(name="微软雅黑", size=11)

    def make_summary(desc_str):
        """从问题描述中提取核心摘要（客户+问题一句话）"""
        if not desc_str or not isinstance(desc_str, str):
            return ""
        import re
        # 去掉【】标记和方括号元数据
        clean = re.sub(r'【[^】]+】', '', desc_str)
        clean = re.sub(r'\[[^\]]+\][^\n]*', '', clean)
        clean = re.sub(r'[\n\r]+', '，', clean)
        clean = re.sub(r'，+', '，', clean)
        clean = re.sub(r'\s+', ' ', clean).strip()
        # 取第一句核心内容，限制60字
        if len(clean) > 60:
            # 尝试在60字内找句号结束
            cut = clean[:60]
            last_period = max(cut.rfind('。'), cut.rfind('，'))
            if last_period > 20:
                clean = clean[:last_period]
            else:
                clean = cut + "..."
        return clean

    def write_bulletin_row(ws, row, text, font_style=None, height=None):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font_style or body_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if height:
            ws.row_dimensions[row].height = height

    r = 1
    write_bulletin_row(ws3, r, f"海外ODM家用空调客诉情况通报（{year}年第{week}周）", title_font, 30)
    r += 1
    write_bulletin_row(ws3, r, f"（{start.strftime('%Y.%m.%d')} — {end.strftime('%Y.%m.%d')}）", body_font, 20)
    r += 2

    overdue_df_for_bulletin = df[
        df['结案状态'].eq('未结案') &
        df['应结案日期'].notna() &
        (df['应结案日期'] < end_dt)
    ]

    sections = [
        ("一、本周新增客诉情况", new_df, "本周新增客诉"),
        ("二、正常跟进中客诉", follow_df, "本周正常跟进客诉"),
        ("三、超期未结案客诉", overdue_df_for_bulletin, "本周超期未结案客诉"),
        ("四、本周结案", closed_df, "本周结案客诉"),
    ]

    for sec_title, sec_df, sec_label in sections:

        count = len(sec_df)
        write_bulletin_row(ws3, r, f"{sec_title}", section_font, 24)
        r += 1
        write_bulletin_row(ws3, r, f"{sec_label}{count}单:", body_font, 20)
        r += 1

        if count == 0:
            write_bulletin_row(ws3, r, "  无。", body_font, 20)
            r += 1
        else:
            for i, (_, rec) in enumerate(sec_df.iterrows()):
                country = str(rec.get("国家或地区", "")) if pd.notna(rec.get("国家或地区", "")) else ""
                desc = make_summary(str(rec.get("问题描述", "")))
                # 避免国家名重复（描述开头可能已含国家名）
                if country and desc.startswith(country):
                    text = "  %d、%s" % (i + 1, desc)
                elif country:
                    text = "  %d、%s%s" % (i + 1, country, desc)
                else:
                    text = "  %d、%s" % (i + 1, desc)
                write_bulletin_row(ws3, r, text, bullet_font, 22)
                r += 1
        r += 1

    ws3.row_dimensions[r].height = 20

    # ============================================
    # Sheet 4 & 5: 赔偿费用 / 材料费用（空模板）
    # ============================================
    for sname, stitle in [("赔偿费用", "赔偿费"), ("材料费用", "售后材料费")]:
        wsx = wb.create_sheet(sname)
        wsx.cell(row=1, column=1, value=stitle).font = Font(name="微软雅黑", size=14, bold=True)
        headers_row = ["序号", "时间", "流程单号", "", "品质流程单号", "客户编码", "分公司", "费用去向", "金额（元）", "是否已转嫁", "跟进人", "责任部门", "备注"]
        for c, h in enumerate(headers_row, 1):
            cell = wsx.cell(row=2, column=c, value=h)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        wsx.cell(row=3, column=13, value="合计").font = Font(name="微软雅黑", size=10)
        wsx.cell(row=3, column=9, value=0).font = Font(name="微软雅黑", size=10)

    return wb


def get_available_weeks(df: pd.DataFrame) -> list:
    """返回数据覆盖的周列表"""
    dates = df["投诉日期"].dropna()
    if len(dates) == 0:
        return []
    min_d = dates.min().date()
    max_d = dates.max().date()
    # 扩展到实际完成日期范围
    if df["实际完成日期"].notna().any():
        max_d2 = df["实际完成日期"].dropna().max().date()
        max_d = max(max_d, max_d2)

    weeks = []
    d = min_d
    while d <= max_d:
        iso = d.isocalendar()
        wk = (iso[0], iso[1])
        if wk not in weeks:
            weeks.append(wk)
        d += timedelta(days=7)
    return sorted(set(weeks))  # unique sorted


if __name__ == "__main__":
    df = load_data()
    # 默认当前周
    today = datetime.now(BJT).date()
    iso = today.isocalendar()
    year, week = iso[0], iso[1]

    if len(sys.argv) >= 3:
        year, week = int(sys.argv[1]), int(sys.argv[2])

    wb = generate_weekly_report(df, year, week)
    out = f"/tmp/海外空调客诉第{week}周周报.xlsx"
    wb.save(out)
    print(f"Generated: {out}")
