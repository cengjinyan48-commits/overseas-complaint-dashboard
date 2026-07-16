#!/usr/bin/env python3
"""
将客诉台账 Excel 转换为天极（FineBI）数据看板可直接导入的标准化数据底表。

输入: 2026年海外客户投诉台账.xlsx (25列，原始台账)
输出: 海外客诉台账_标准化数据.xlsx (33列，含8个计算字段 + 数据字典)

用法:
    python convert_for_taiji.py                        # 默认路径
    python convert_for_taiji.py --input 台账.xlsx       # 指定输入
    python convert_for_taiji.py --output 天极数据.xlsx  # 指定输出
"""

import argparse
import os
import sys
from datetime import datetime, date

import openpyxl


# ── 输入文件的列结构（客诉台账，25列）─────────────────────────────────────
SRC_HEADERS = [
    "编号", "分公司", "国家或地区", "是否大客户", "投诉日期",
    "应结案日期", "结案预警", "实际完成日期", "完成周期（天）", "机型属性",
    "故障比例", "问题描述", "客户诉求", "跟进人", "处理类型",
    "结案状态", "应急措施", "原因分析", "长期整改措施", "责任单位",
    "故障大类", "品质负责人", "8D报告", "8D措施点检", "备注",
]

# ── 输出文件比输入多出的 8 个计算列 ─────────────────────────────────────
COMPUTED_HEADERS = [
    "投诉年月", "投诉月份", "是否超期", "是否结案",
    "是否有8D", "结案标记", "未结案标记", "超期标记",
]

ALL_HEADERS = SRC_HEADERS + COMPUTED_HEADERS
SRC_COL_COUNT = len(SRC_HEADERS)  # 25
OUT_COL_COUNT = len(ALL_HEADERS)  # 33


def _is_empty_row(row_data: list) -> bool:
    """判断是否为空行/模板行：编号为空或只有编号+公式残留"""
    if row_data[0] is None:
        return True
    non_empty = sum(1 for v in row_data if v is not None)
    return non_empty <= 2


def convert(source_path: str, output_path: str) -> int:
    """
    执行转换，返回有效数据行数。

    Parameters
    ----------
    source_path : str
        客诉台账 Excel 文件路径
    output_path : str
        输出的标准化数据底表路径

    Returns
    -------
    int
        有效数据行数
    """
    # 用 data_only=True 读取，确保公式单元格拿到缓存值
    src = openpyxl.load_workbook(source_path, data_only=True)
    ws_src = src["所有客诉"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客诉明细"

    # ── 写入表头 ──────────────────────────────────────────────────
    for i, h in enumerate(ALL_HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    today = date.today()
    data_count = 0
    skipped = 0

    # ── 遍历数据行 ─────────────────────────────────────────────────
    for r in range(3, ws_src.max_row + 1):
        row_data = [ws_src.cell(row=r, column=c).value for c in range(1, SRC_COL_COUNT + 1)]

        if _is_empty_row(row_data):
            skipped += 1
            continue

        data_count += 1
        out_row = data_count + 1  # +1 for header

        # 写入原始列
        for c in range(1, SRC_COL_COUNT + 1):
            ws.cell(row=out_row, column=c, value=row_data[c - 1])

        # 提取关键字段
        投诉日期 = row_data[4]   # 列 E
        应结案日期 = row_data[5]  # 列 F
        结案状态 = row_data[15]   # 列 P
        八D报告 = row_data[22]    # 列 W

        # ── 计算派生字段 ──────────────────────────────────────────

        # 26: 投诉年月
        if 投诉日期 and isinstance(投诉日期, (datetime, date)):
            投诉年月 = 投诉日期.strftime("%Y-%m") if isinstance(投诉日期, datetime) else 投诉日期.strftime("%Y-%m")
            投诉月份 = 投诉日期.month
        else:
            投诉年月 = None
            投诉月份 = None

        # 29: 是否结案
        是否结案 = 结案状态 in ("结案", "关闭")

        # 31: 结案标记
        结案标记 = 1 if 是否结案 else 0

        # 32: 未结案标记（未结案 + 暂停）
        未结案标记 = 0 if 是否结案 else 1

        # 28: 是否超期
        if 未结案标记 == 1 and 应结案日期 and isinstance(应结案日期, (datetime, date)):
            应结案日期_d = 应结案日期.date() if isinstance(应结案日期, datetime) else 应结案日期
            是否超期 = 应结案日期_d < today
        else:
            是否超期 = False

        # 33: 超期标记
        超期标记 = 1 if (未结案标记 == 1 and 是否超期) else 0

        # 30: 是否有8D
        是否有8D = (八D报告 == "有")

        # 写入计算列
        ws.cell(row=out_row, column=26, value=投诉年月)
        ws.cell(row=out_row, column=27, value=投诉月份)
        ws.cell(row=out_row, column=28, value=是否超期)
        ws.cell(row=out_row, column=29, value=是否结案)
        ws.cell(row=out_row, column=30, value=是否有8D)
        ws.cell(row=out_row, column=31, value=结案标记)
        ws.cell(row=out_row, column=32, value=未结案标记)
        ws.cell(row=out_row, column=33, value=超期标记)

    # ── 数据字典 Sheet ────────────────────────────────────────────
    ws_dict = wb.create_sheet("数据字典")
    dict_headers = ["字段名", "数据类型", "非空数", "说明"]
    dict_rows = [
        ["编号", "int64", data_count, "客诉编号"],
        ["分公司", "object", None, "海外分公司"],
        ["国家或地区", "object", None, "国家/地区"],
        ["是否大客户", "object", None, "是否为大客户"],
        ["投诉日期", "datetime64[ns]", None, "投诉受理日期"],
        ["应结案日期", "datetime64[ns]", None, "应完成结案日期"],
        ["结案预警", "datetime64[ns]", None, "超期预警日期"],
        ["实际完成日期", "datetime64[ns]", None, "实际完成日期"],
        ["完成周期（天）", "float64", None, "从投诉到结案的天数"],
        ["机型属性", "object", None, "产品机型属性"],
        ["故障比例", "object", None, "故障比例描述"],
        ["问题描述", "object", None, "客户问题详细描述"],
        ["客户诉求", "object", None, "客户具体诉求"],
        ["跟进人", "object", None, "跟进处理人员"],
        ["处理类型", "object", None, "处理类型（客服/品质）"],
        ["结案状态", "object", None, "当前结案状态"],
        ["应急措施", "object", None, "应急处理措施"],
        ["原因分析", "object", None, "根因分析"],
        ["长期整改措施", "object", None, "长期整改方案"],
        ["责任单位", "object", None, "责任归属单位"],
        ["故障大类", "object", None, "故障分类大类"],
        ["品质负责人", "object", None, "品质负责人员"],
        ["8D报告", "object", None, "是否有8D报告"],
        ["8D措施点检", "object", None, "8D措施执行点检"],
        ["备注", "object", None, "备注信息"],
        ["投诉年月", "object", None, "投诉年月（YYYY-MM）"],
        ["投诉月份", "int32", None, "投诉月份（1-12）"],
        ["是否超期", "bool", None, "是否超期标记"],
        ["是否结案", "bool", None, "是否结案标记"],
        ["是否有8D", "bool", None, "是否有8D标记"],
        ["结案标记", "int64", None, "1=结案+关闭, 0=其他"],
        ["未结案标记", "int64", None, "1=未结案+暂停, 0=结案+关闭"],
        ["超期标记", "int64", None, "1=超期, 0=未超期"],
    ]

    for i, h in enumerate(dict_headers, 1):
        cell = ws_dict.cell(row=1, column=i, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
    for r, row_data in enumerate(dict_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws_dict.cell(row=r, column=c, value=val)

    # ── 保存 ──────────────────────────────────────────────────────
    wb.save(output_path)

    # ── 统计输出 ──────────────────────────────────────────────────
    结案标 = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=31).value == 1)
    未结案标 = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=32).value == 1)
    超期标 = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=33).value == 1)
    八个D = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=30).value is True)

    print(f"转换完成: {data_count} 条数据, 跳过 {skipped} 个空行")
    print(f"结案标记={结案标}, 未结案标记={未结案标}, 超期标记={超期标}, 有8D={八个D}")
    print(f"结案率 = {结案标}/{data_count} = {结案标 / data_count * 100:.1f}%")
    print(f"8D覆盖率 = {八个D}/{data_count} = {八个D / data_count * 100:.1f}%")
    print(f"输出: {output_path}")

    return data_count


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    default_input = os.path.join(script_dir, "2026年海外客户投诉台账.xlsx")
    default_output = os.path.join(script_dir, "海外客诉台账_标准化数据.xlsx")

    parser = argparse.ArgumentParser(description="客诉台账 → 天极标准化数据底表")
    parser.add_argument("--input", "-i", default=default_input, help="输入的客诉台账 Excel 路径")
    parser.add_argument("--output", "-o", default=default_output, help="输出的标准化数据底表路径")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"错误: 找不到输入文件 {args.input}", file=sys.stderr)
        sys.exit(1)

    convert(args.input, args.output)


if __name__ == "__main__":
    main()